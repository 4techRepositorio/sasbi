"""Extracção tabular para persistência e motor de query (TICKET-016)."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from fourpro_api.models.ingestion import PARSED_ROWS_CAP

_MAX_TEXT_SCAN = 20_000_000


def extract_tabular_rows(
    path: Path,
    *,
    body: bytes | None = None,
    cap: int = PARSED_ROWS_CAP,
) -> tuple[list[dict[str, Any]], str, bool]:
    """Devolve (rows, summary, truncated).

    Cap documentado: no máximo ``PARSED_ROWS_CAP`` (50 000) linhas em BD.
    """
    ext = path.suffix.lower().lstrip(".")
    raw = body if body is not None else path.read_bytes()

    if ext in ("csv", "txt"):
        return _from_delimited(raw, ext=ext, cap=cap)
    if ext == "json":
        return _from_json(raw, cap=cap)
    if ext == "xlsx":
        return _from_xlsx(path, cap=cap)
    if ext == "xls":
        return _from_xls(path, cap=cap)
    raise ValueError(f"tipo não suportado: {ext}")


def _cell_value(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _from_delimited(raw: bytes, *, ext: str, cap: int) -> tuple[list[dict[str, Any]], str, bool]:
    text = raw.decode("utf-8", errors="replace")
    if len(text) > _MAX_TEXT_SCAN:
        text = text[:_MAX_TEXT_SCAN]
    if ext == "txt":
        lines = [ln for ln in text.splitlines() if ln.strip()]
        rows = [{"line": ln, "line_no": i + 1} for i, ln in enumerate(lines[:cap])]
        truncated = len(lines) > cap
        summary = f"txt_len={len(text)};rows_stored={len(rows)}"
        if truncated:
            summary += f";truncated_at={cap}"
        return rows, summary, truncated

    reader = csv.DictReader(text.splitlines())
    rows: list[dict[str, Any]] = []
    truncated = False
    for i, row in enumerate(reader):
        if i >= cap:
            truncated = True
            break
        rows.append({k: _cell_value(v) for k, v in row.items() if k is not None})
    summary = f"csv_rows_stored={len(rows)}"
    if truncated:
        summary += f";truncated_at={cap}"
    return rows, summary, truncated


def _from_json(raw: bytes, *, cap: int) -> tuple[list[dict[str, Any]], str, bool]:
    data = json.loads(raw.decode("utf-8", errors="strict"))
    if isinstance(data, list):
        items = data
        if items and all(isinstance(x, dict) for x in items):
            rows = [{str(k): _cell_value(v) for k, v in item.items()} for item in items[:cap]]
            truncated = len(items) > cap
            summary = f"json_list_len={len(items)};rows_stored={len(rows)}"
            if truncated:
                summary += f";truncated_at={cap}"
            return rows, summary, truncated
        rows = [{"value": _cell_value(x), "index": i} for i, x in enumerate(items[:cap])]
        truncated = len(items) > cap
        summary = f"json_list_len={len(items)};rows_stored={len(rows)}"
        if truncated:
            summary += f";truncated_at={cap}"
        return rows, summary, truncated
    if isinstance(data, dict):
        rows = [{"key": str(k), "value": _cell_value(v)} for k, v in list(data.items())[:cap]]
        truncated = len(data) > cap
        summary = f"json_keys={len(data)};rows_stored={len(rows)}"
        if truncated:
            summary += f";truncated_at={cap}"
        return rows, summary, truncated
    return [{"value": _cell_value(data)}], "json_scalar;rows_stored=1", False


def _from_xlsx(path: Path, *, cap: int) -> tuple[list[dict[str, Any]], str, bool]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb.active
        title = ws.title or "sheet1"
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return [], f"xlsx_active_sheet={title}_rows_stored=0", False
        headers = [
            str(h).strip() if h is not None and str(h).strip() else f"col_{i}"
            for i, h in enumerate(header_row)
        ]
        rows: list[dict[str, Any]] = []
        truncated = False
        for values in it:
            if len(rows) >= cap:
                truncated = True
                break
            if values is None or all(v is None for v in values):
                continue
            row = {
                headers[i]: _cell_value(values[i] if i < len(values) else None)
                for i in range(len(headers))
            }
            rows.append(row)
        summary = f"xlsx_active_sheet={title}_rows_stored={len(rows)}"
        if truncated:
            summary += f";truncated_at={cap}"
        return rows, summary, truncated
    finally:
        wb.close()


def _from_xls(path: Path, *, cap: int) -> tuple[list[dict[str, Any]], str, bool]:
    import xlrd

    book = xlrd.open_workbook(path, on_demand=True)
    try:
        sh = book.sheet_by_index(0)
        if sh.nrows == 0:
            return [], "xls_sheet0_rows_stored=0", False
        headers = [
            str(sh.cell_value(0, c)).strip() or f"col_{c}" for c in range(sh.ncols)
        ]
        data_rows = sh.nrows - 1
        take = min(data_rows, cap)
        rows: list[dict[str, Any]] = []
        for r in range(1, take + 1):
            rows.append(
                {headers[c]: _cell_value(sh.cell_value(r, c)) for c in range(sh.ncols)}
            )
        truncated = data_rows > cap
        summary = f"xls_sheet0_rows_stored={len(rows)}"
        if truncated:
            summary += f";truncated_at={cap}"
        return rows, summary, truncated
    finally:
        book.release_resources()
