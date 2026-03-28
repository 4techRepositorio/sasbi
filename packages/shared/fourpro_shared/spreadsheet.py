"""Métricas leves de folhas de cálculo (API + worker)."""

from __future__ import annotations

from pathlib import Path


class SpreadsheetSummaryError(Exception):
    """Não foi possível resumir o ficheiro."""


def summarize_workbook(path: Path) -> str:
    """Devolve texto curto com contagens (primeira folha ou todas, conforme tipo)."""
    ext = path.suffix.lower().lstrip(".")
    if ext == "xlsx":
        return _summarize_xlsx(path)
    if ext == "xls":
        return _summarize_xls(path)
    raise SpreadsheetSummaryError(f"extensão não suportada para workbook: {ext}")


def _summarize_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        ws = wb.active
        title = ws.title or "sheet1"
        max_scan = 500_000
        n = 0
        for _ in ws.iter_rows():
            n += 1
            if n >= max_scan:
                break
        truncated = "+" if n >= max_scan else ""
        return f"xlsx_active_sheet={title}_rows_est={n}{truncated}"
    finally:
        wb.close()


def _summarize_xls(path: Path) -> str:
    import xlrd

    book = xlrd.open_workbook(path, on_demand=True)
    try:
        sh = book.sheet_by_index(0)
        return f"xls_sheet0_rows={sh.nrows}_cols={sh.ncols}"
    finally:
        book.release_resources()
